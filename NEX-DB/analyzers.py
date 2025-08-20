# analyzers.py
import pandas as pd
import re
from abc import ABC, abstractmethod
from datetime import datetime
from openpyxl import load_workbook
import dateparser
UNK_TOKENS = {"UNK", "???", "###", "N/A", "NA", "-", "NULL", "？", "؟", ""}

def cell_ref(row_idx: int, col_idx: int) -> str:
    col_letter = chr(65 + col_idx)
    return f"{col_letter}{row_idx + 2}"

def detect_and_parse_dates(df: pd.DataFrame) -> pd.DataFrame:
    date_regexes = [
        r'\b\d{1,2}[/-]\d{1,2}[/-]\d{2,4}\b',
    ]

    months_en = (
        r'(?:January|February|March|April|May|June|July|August|'
        r'September|October|November|December)'
    )
    months_ar = (
        r'(?:يناير|فبراير|مارس|أبريل|مايو|يونيو|يوليو|أغسطس|'
        r'سبتمبر|أكتوبر|نوفمبر|ديسمبر)'
    )
    date_regexes += [
        fr'\b\d{{1,2}}\s+{months_en}\s+\d{{4}}\b',
        fr'\b\d{{1,2}}\s+{months_ar}\s+\d{{4}}\b',
    ]

    weekdays_en = r'(?:Monday|Tuesday|Wednesday|Thursday|Friday|Saturday|Sunday)'
    weekdays_ar = r'(?:الاثنين|الثلاثاء|الأربعاء|الخميس|الجمعة|السبت|الأحد)'
    date_regexes += [
        fr'\b{weekdays_en}\s+{months_en}\s+\d{{1,2}},?\s*\d{{4}}\b',
        fr'\b{weekdays_ar}\s+{months_ar}\s+\d{{1,2}},?\s*\d{{4}}\b',
    ]
    combined_pattern = re.compile('|'.join(date_regexes), flags=re.IGNORECASE)
    for col in df.columns:
        if df[col].dtype == object:
            def parse_cell(val):
                if not isinstance(val, str):
                    return val
                m = combined_pattern.search(val)
                if m:
                    return dateparser.parse(m.group(), languages=['en', 'ar'])
                return val
            df[col] = df[col].map(parse_cell)
    for col in df.columns:
        df[col] = pd.to_datetime(df[col], errors='ignore', dayfirst=True)
    return df

class BaseAnalyzer(ABC):
    @abstractmethod
    def run(self, df: pd.DataFrame, **kwargs) -> list[dict]:
        pass

ANALYZERS: list[BaseAnalyzer] = []

def register(cls: type[BaseAnalyzer]) -> type[BaseAnalyzer]:
    ANALYZERS.append(cls())
    return cls


def cell_ref(row_idx: int, col_idx: int) -> str:
    col_letter = chr(65 + col_idx)
    return f"{col_letter}{row_idx + 2}"  


class BaseAnalyzer(ABC):
    @abstractmethod
    def run(self, df: pd.DataFrame, **kwargs) -> list[dict]:
        pass


ANALYZERS: list[type[BaseAnalyzer]] = []
def register(cls: type[BaseAnalyzer]) -> type[BaseAnalyzer]:
    ANALYZERS.append(cls)
    return cls


@register
class MissingDataAnalyzer(BaseAnalyzer):
    def run(self, df: pd.DataFrame, **kwargs) -> list[dict]:
        issues = []
        n_rows, n_cols = df.shape
        for c_idx, col in enumerate(df.columns):
            if df[col].isna().all():
                issues.append({
                    "column": col,
                    "issue": "All values Missing On Column",
                    "count": n_rows,
                    "pct": "100%",
                    "details": "Column is Empty",
                    "rows": "-"
                })
        full_missing = df.isna().all(axis=1)
        rows = [cell_ref(i, 0) + f":{cell_ref(i, n_cols-1)}" 
                for i in full_missing[full_missing].index]
        if rows:
            issues.append({
                "column": "ALL",
                "issue": "Missing Row",
                "count": len(rows),
                "pct": f"{int(len(rows)/n_rows*100)}%",
                "details": "entire rows missing",
                "rows": ", ".join(rows)
            })

        return issues

@register
class DuplicateDataAnalyzer(BaseAnalyzer):
    def run(self, df: pd.DataFrame, **kwargs) -> list[dict]:
        issues = []
        n_rows, _ = df.shape
        dup_mask = df.duplicated(keep=False)
        dup_rows = [cell_ref(i, 0) + f":{cell_ref(i, df.shape[1]-1)}"
                    for i in dup_mask[dup_mask].index]
        if dup_rows:
            issues.append({
                "column": "ALL",
                "issue": "Full Duplicate Rows",
                "count": len(dup_rows),
                "pct": f"{int(len(dup_rows)/n_rows*100)}%",
                "details": "identical rows",
                "rows": ", ".join(dup_rows[:10]) + ("..." if len(dup_rows) > 10 else "")
            })
        return issues


@register
class InvalidValuesAnalyzer(BaseAnalyzer):
    def run(self, df: pd.DataFrame, **kwargs) -> list[dict]:
        issues = []
        n_rows, _ = df.shape
        for c_idx, col in enumerate(df.columns):
            if pd.api.types.is_numeric_dtype(df[col]):
                series = df[col]
                neg = series < 0
                if neg.any():
                    rows = [cell_ref(i, c_idx) for i in series[neg].index]
                    issues.append({
                        "column": col,
                        "issue": "Negative Values",
                        "count": int(neg.sum()),
                        "pct": f"{int(neg.sum()/n_rows*100)}%",
                        "details": "negative not allowed",
                        "rows": ", ".join(rows[:10]) + ("..." if len(rows) > 10 else "")
                    })
                zero = series == 0
                if zero.any():
                    rows = [cell_ref(i, c_idx) for i in series[zero].index]
                    issues.append({
                        "column": col,
                        "issue": "Zero Values",
                        "count": int(zero.sum()),
                        "pct": f"{int(zero.sum()/n_rows*100)}%",
                        "details": "zero may be invalid",
                        "rows": ", ".join(rows[:10]) + ("..." if len(rows) > 10 else "")
                    })
        return issues
    
@register
class OutliersAnalyzer(BaseAnalyzer):
    def run(self, df: pd.DataFrame, **kwargs) -> list[dict]:
        issues = []
        n_rows, _ = df.shape
        for c_idx, col in enumerate(df.columns):
            if pd.api.types.is_numeric_dtype(df[col]):
                series = df[col].dropna()
                mean, std = series.mean(), series.std()
                mask = (series < mean - 3*std) | (series > mean + 3*std)
                if mask.any():
                    rows = [cell_ref(i, c_idx) for i in series[mask].index]
                    issues.append({
                        "column": col,
                        "issue": "Outliers",
                        "count": int(mask.sum()),
                        "pct": f"{int(mask.sum()/n_rows*100)}%",
                        "details": f"outside ±3σ (mean={mean:.2f})",
                        "rows": "-"
                    })
        return issues

@register
class ColumnSimilarityAnalyzer(BaseAnalyzer):
    def run(self, df: pd.DataFrame, **kwargs) -> list[dict]:
        issues = []
        threshold = kwargs.get("similarity_threshold", 0.8)
        n_rows, _ = df.shape
        for i, col1 in enumerate(df.columns):
            for j, col2 in enumerate(df.columns):
                if i >= j:
                    continue 
                series1 = df[col1].dropna()
                series2 = df[col2].dropna()
                min_len = min(len(series1), len(series2))
                if min_len == 0:
                    continue
                series1 = series1.iloc[:min_len].astype(str).reset_index(drop=True)
                series2 = series2.iloc[:min_len].astype(str).reset_index(drop=True)
                match_count = (series1 == series2).sum()
                similarity = match_count / min_len
                if similarity >= threshold:
                    issues.append({
                        "column": col1,
                        "issue": "There Are Some Columns Match",
                        "count": match_count,
                        "pct": f"{int(similarity * 100)}%",
                        "details": f"The column Similar to '{col2}'",
                        "rows": "-"
                    })
        return issues

@register
class CrossFieldValueAnalyzer(BaseAnalyzer):
    def run(self, df: pd.DataFrame, rules: list = None, **kwargs) -> list[dict]:
        from collections import defaultdict
        issues_map = defaultdict(lambda: {
            "count": 0,
            "rows": [],
            "details": "",
            "pct": "",
            "column": "",
            "issue": ""
        })
        n_rows, _ = df.shape
        date_columns = [col for col in df.columns if pd.api.types.is_datetime64_any_dtype(df[col])]
        if rules is None:
            rules = [
                lambda row: self.rule_start_before_end(row, date_columns),
                self.rule_zero_qty_price,
                self.rule_country_currency,
                self.rule_gender_name
            ]
        for i in range(n_rows):
            row = df.iloc[i]

            for rule in rules:
                try:
                    results = rule(row)
                    for result in results:
                        col_name = result.get("column", "—")
                        issue_text = result.get("issue", "Cross-field Value Error")
                        details = result.get("details", "")
                        col_idx = df.columns.get_loc(col_name) if col_name in df.columns else None
                        cell_ref = self.cell_ref(i, col_idx) if col_idx is not None else f"Row {i + 2}"

                        key = (col_name, issue_text)
                        issues_map[key]["column"] = col_name
                        issues_map[key]["issue"] = issue_text
                        issues_map[key]["details"] = details
                        issues_map[key]["count"] += 1
                        issues_map[key]["rows"].append(cell_ref)
                except:
                    continue
        issues = []
        for issue in issues_map.values():
            issue["pct"] = f"{int(100 * issue['count'] / n_rows)}%"
            issue["rows"] = ", ".join(issue["rows"])
            issues.append(issue)
        return issues
    @staticmethod
    def cell_ref(row_idx: int, col_idx: int) -> str:
        col_letter = chr(65 + col_idx)
        return f"{col_letter}{row_idx + 2}"

    def rule_start_before_end(self, row, date_columns):
        if len(date_columns) >= 2:
            try:
                start = pd.to_datetime(row[date_columns[0]], errors='coerce')
                end = pd.to_datetime(row[date_columns[1]], errors='coerce')
                if pd.notnull(start) and pd.notnull(end) and start > end:
                    return [{
                        "column": date_columns[0],
                        "issue": "Start date is after end date",
                        "details": f"{date_columns[0]} > {date_columns[1]}"
                    }]
            except:
                pass
        return []
    @staticmethod
    def rule_zero_qty_price(row):
        try:
            qty = float(str(row["Quantity"]).replace(",", ""))
            price = float(str(row["Price"]).replace(",", ""))
            if qty == 0 and price != 0:
                return [{
                    "column": "Price",
                    "issue": "Zero quantity with non-zero price",
                    "details": f"Quantity=0, Price={price}"
                }]
        except:
            pass
        return []
    @staticmethod
    def rule_country_currency(row):
        country = str(row.get("Country", "")).strip().lower()
        currency = str(row.get("Currency", "")).strip().lower()
        expected_currency = {
            "egypt": "ج.م",
            "usa": "$",
            "uk": "£",
            "ksa": "ر.س"
        }
        if country in expected_currency:
            if expected_currency[country].lower() not in currency:
                return [{
                    "column": "Currency",
                    "issue": "Currency mismatch",
                    "details": f"Expected {expected_currency[country]} for {country.title()}, got {currency}"
                }]
        return []
    @staticmethod
    def rule_gender_name(row):
        name = str(row.get("Name", "")).strip().lower()
        gender = str(row.get("Gender", "")).strip().lower()
        female_names = {"fatima", "sara", "laila", "eman", "nour"}
        if gender == "male" and any(fname in name for fname in female_names):
            return [{
                "column": "Gender",
                "issue": "Male gender with female name",
                "details": f"Name={name}, Gender={gender}"
            }]
        return []

@register
class MixedTypeAnalyzer(BaseAnalyzer):
    currency_pattern = re.compile(r'[\d,.]+\s*(\$|€|£|ج\.م|د\.ك|ر\.س|AED|SAR)', re.IGNORECASE)
    percentage_pattern = re.compile(r'\d+(\.\d+)?\s*(%|٪)')
    unit_pattern = re.compile(r'\d+(\.\d+)?\s*(kg|g|mg|lb|m|cm|mm|km|ltr|ml)', re.IGNORECASE)
    boolean_values = {"true", "false", "yes", "no", "نعم", "لا"}
    def detect_type(self, value: str) -> str:
        if pd.isna(value) or str(value).strip() == "":
            return "empty"
        val = str(value).strip().lower()
        if val in self.boolean_values:
            return "boolean"
        if self.currency_pattern.search(val):
            return "currency"
        if self.percentage_pattern.search(val):
            return "percentage"
        if self.unit_pattern.search(val):
            return "unit"
        try:
            pd.to_datetime(value, errors='raise')
            return "date"
        except:
            pass
        try:
            float(value)
            return "number"
        except:
            pass
        return "text"
    def run(self, df: pd.DataFrame, clean: bool = False, **kwargs) -> dict | list[dict]:
        issues = []
        n_rows, _ = df.shape
        cleaned_df = df.copy()
        for c_idx, col in enumerate(df.columns):
            series = df[col].fillna("").astype(str)
            type_series = series.apply(self.detect_type)
            type_counts = type_series.value_counts()
            dominant_type = type_counts.idxmax()
            mixed_rows = [
                self.cell_ref(i, c_idx)
                for i, t in enumerate(type_series)
                if t != dominant_type
            ]
            if len(type_counts) > 1:
                issues.append({
                    "column": col,
                    "issue": "Mixed Data Types",
                    "count": len(mixed_rows),
                    "pct": f"{int(len(mixed_rows)/n_rows*100)}%",
                    "details": f"Dominant type: {dominant_type}, others: {', '.join([t for t in type_counts.index if t != dominant_type])}",
                    "distribution": type_counts.to_dict(),
                    "rows": ", ".join(mixed_rows)
                })
                if clean:
                    cleaned_df[col] = [
                        val if self.detect_type(val) == dominant_type else None
                        for val in df[col]
                    ]
            empty_count = type_counts.get("empty", 0)
            empty_pct = empty_count / n_rows
            if empty_pct > 0.8:
                empty_rows = [
                    self.cell_ref(i, c_idx)
                    for i, t in enumerate(type_series)
                    if t == "empty"
                ]
                issues.append({
                    "column": col,
                    "issue": "Mostly Empty Column",
                    "count": empty_count,
                    "pct": f"{int(empty_pct * 100)}%",
                    "details": "Column contains mostly empty or missing values",
                    "rows": ", ".join(empty_rows)
                })
        if clean:
            return {"issues": issues, "cleaned_df": cleaned_df}
        else:
            return issues
    @staticmethod
    def cell_ref(row_idx: int, col_idx: int) -> str:
        col_letter = chr(65 + col_idx)
        return f"{col_letter}{row_idx + 2}"
@register
class TemporalErrorsAnalyzer(BaseAnalyzer):
    def run(self, df: pd.DataFrame, **kwargs) -> list[dict]:
        issues = []
        def col_idx_to_excel_letter(idx: int) -> str:
            letters = ""
            while idx >= 0:
                idx, remainder = divmod(idx, 26)
                letters = chr(65 + remainder) + letters
                idx -= 1
            return letters
        for c_idx, col in enumerate(df.columns):
            if 'date' in col.lower():
                series = pd.to_datetime(df[col], errors='coerce')
                if not series.is_monotonic_increasing:
                    violations_mask = series.diff() < pd.Timedelta(0)
                    violations_count = violations_mask.sum()
                    pct = f"{(violations_count / series.size * 100):.2f}%"
                    rows_list = df[violations_mask].index.tolist()
                    cell_refs = [f"{col_idx_to_excel_letter(c_idx)}{row + 2}" for row in rows_list]
                    rows_str = ", ".join(cell_refs) if cell_refs else "-"
                    issues.append({
                        "column": col,
                        "issue": "Time Repetition Error",
                        "count": int(violations_count),
                        "pct": pct,
                        "details": "Dates not in chronological order",
                        "rows": rows_str
                    })
        return issues
@register
class InvalidDateValuesAnalyzer(BaseAnalyzer):
    def run(self, df: pd.DataFrame, **kwargs) -> list[dict]:
        issues = []
        keywords = kwargs.get("keywords", ["خطأ", "غير معروف", "n/a", "unknown", "NULL", "null", "#", "N/A", "NaT", "nat", "NAT","?","؟","#DIV/0!", "#REF!", "#VALUE!", "#NAME?", "#NULL!", "#NUM!", "#N/A"])
        for c_idx, col in enumerate(df.columns):
            null_mask = df[col].isna()
            if null_mask.any():
                rows = [cell_ref(i, c_idx) for i in null_mask[null_mask].index]
                issues.append({
                    "column": col,
                    "issue": "Missing values",
                    "count": len(rows),
                    "pct": f"{(len(rows) / len(df) * 100):.2f}%",
                    "details": "Null values or Excel Error",
                    "rows": ", ".join(rows[:10]) + ("..." if len(rows) > 10 else "")
                })
            str_col = df[col].astype(str).str.lower()
            keywords_lower = [kw.lower() for kw in keywords]
            keyword_mask = str_col.isin(keywords_lower)
            if keyword_mask.any():
                matched_values = str_col[keyword_mask].unique()
                rows = [cell_ref(i, c_idx) for i in keyword_mask[keyword_mask].index]
                issues.append({
                    "column": col,
                    "issue": "Found Unacceptable Keyword",
                    "count": len(rows),
                    "pct": f"{(len(rows) / len(df) * 100):.2f}%",
                    "details": f"Found keywords: {', '.join(sorted(set(matched_values)))}",
                    "rows": ", ".join(rows[:10]) + ("..." if len(rows) > 10 else "")
                })
        return issues
@register
class InvalidDateFormatAnalyzer(BaseAnalyzer):
    default_formats = [
        "%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y",
        "%Y/%m/%d", "%d-%m-%Y", "%d.%m.%Y",
        "%Y-%m-%d %H:%M:%S", "%d/%m/%Y %H:%M",
        "%d-%b-%Y", "%d-%b-%y", "%b %d, %Y",
        "%Y.%m.%d", "%Y%m%d", "%d/%m/%y",
        "%Y-%m-%dT%H:%M:%S", "%d-%m-%Y %H:%M",
        "%Y/%m/%d %I:%M %p"
    ]
    @staticmethod
    def cell_ref(row_idx: int, col_idx: int) -> str:
        col_letter = chr(65 + col_idx)
        return f"{col_letter}{row_idx + 2}"
    def run(
        self,
        df: pd.DataFrame,
        excel_file: str = None,
        sheet_name: str = None,
        column_types: dict[str, str] = None,
        valid_formats: list[str] = None,
        **kwargs
    ) -> list[dict]:
        issues = []
        n_rows = len(df)
        cols = list(df.columns)
        valid_formats = valid_formats or self.default_formats
        col_types: dict[str, str] = {}
        if column_types:
            col_types = column_types.copy()
        elif excel_file and sheet_name:
            wb = load_workbook(excel_file, data_only=True)
            ws = wb[sheet_name]
            for idx, col in enumerate(cols, start=1):
                cell = ws.cell(row=2, column=idx)
                fmt = (cell.number_format or "").lower()
                dtype = cell.data_type
                if any(x in fmt for x in ["y", "m", "d"]) or dtype == "d":
                    col_types[col] = "date"
                else:
                    col_types[col] = "other"
        else:
            for col in cols:
                col_types[col] = "date" if "date" in col.lower() else "other"
        for c_idx, col in enumerate(cols):
            if col_types.get(col) != "date":
                continue
            raw = df[col].astype(str).str.strip()
            raw = raw.str.replace(r"[^\w\s/:.\-]", "", regex=True).str.strip()
            raw = raw[raw.str.strip().str.lower().isin(["", "nan", "nat", "none"]) == False]
            failed = pd.Series([True] * len(raw), index=raw.index)
            parsed = pd.to_datetime(raw, errors="coerce", infer_datetime_format=False)
            failed &= parsed.isna()
            for idx in raw[failed].index:
                val = raw.loc[idx]
                for fmt in valid_formats:
                    try:
                        datetime.strptime(val, fmt)
                        failed.loc[idx] = False
                        break
                    except ValueError:
                        continue
            final_failed = failed[failed].index
            if final_failed.any():
                rows = [self.cell_ref(i, c_idx) for i in final_failed]
                examples = [raw.loc[i] for i in final_failed[:3]]
                pct = len(final_failed) / n_rows * 100
                issues.append({
                    "column": col,
                    "issue": "Invalid Date Format",
                    "count": len(final_failed),
                    "pct": f"{pct:.2f}%",
                    "details": (
                        f"Unrecognized date in rows "
                        f"{', '.join(map(str, final_failed[:3]))}. "
                        f"Examples: {', '.join(examples)}"
                    ),
                    "rows": ", ".join(rows[:10]) + ("..." if len(rows) > 10 else "")
                })
        return issues
@register
class DecimalFormatAnalyzer(BaseAnalyzer):
    def run(self, df: pd.DataFrame, **kwargs) -> list[dict]:
        import re
        from dateutil.parser import parse
        issues = []
        valid_number_pattern = r'^-?\d{1,3}(,\d{3})*(\.\d+)?$|^-?\d+(\.\d+)?$'
        for c_idx, col in enumerate(df.columns):
            str_col = df[col].astype(str).str.strip()
            suspicious = []

            for i, val in str_col.items():
                val_clean = val.replace(" ", "")
                if val_clean == "":
                    continue
                try:
                    parse(val, fuzzy=False)
                    continue
                except:
                    pass
                if any(char.isdigit() for char in val_clean):
                    if re.fullmatch(valid_number_pattern, val_clean):
                        continue 
                    suspicious.append((i, val))
            if suspicious:
                rows = [self.cell_ref(i, c_idx) for i, _ in suspicious]
                values = [v for _, v in suspicious]

                issues.append({
                    "column": col,
                    "issue": "There Are Symbols In Cells",
                    "count": len(rows),
                    "pct": f"{(len(rows) / len(df) * 100):.2f}%",
                    "details": f"Invalid values: {', '.join(values[:3])}" + ("..." if len(values) > 3 else ""),
                    "rows": ", ".join(rows)
                })
        return issues
    @staticmethod
    def cell_ref(row_idx: int, col_idx: int) -> str:
        col_letter = chr(65 + col_idx)
        return f"{col_letter}{row_idx + 2}"
def run_all(df: pd.DataFrame, **kwargs) -> list[dict]:
    results = []
    for Analyzer in ANALYZERS:
        results.extend(Analyzer().run(df, **kwargs))
    return results


