# relationships.py
from difflib import SequenceMatcher
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import pandas as pd

def compute_similarity(a: str, b: str) -> float:

    return SequenceMatcher(None, a.lower(), b.lower()).ratio()

def compute_relationships(
    file_dfs: dict[str, pd.DataFrame],
    central_files: list[str],
    threshold: float = 0.9
) -> list[dict]:

    relationships = []

    existing = [f for f in central_files if f in file_dfs]
    missing = set(central_files) - set(existing)
    for m in missing:
        print(f"⚠️ Warning: central key file '{m}' not found—skipping it.")

    for file_a in existing:
        cols_a = file_dfs[file_a].columns
        for col_a in cols_a:
            for file_b, df_b in file_dfs.items():
                if file_b == file_a:
                    continue
                for col_b in df_b.columns:
                    sim = compute_similarity(col_a, col_b)
                    if sim >= threshold:
                        relationships.append({
                            "file_a":   file_a,
                            "column_a": col_a,
                            "file_b":   file_b,
                            "column_b": col_b,
                            "rating":   f"{sim * 100:.2f}%"
                        })
    return relationships

def add_relationships_to_report(report_path: str, relationships: list[dict]):

    wb = load_workbook(report_path)
    ws = wb.create_sheet("Relationships")


    side   = Side(border_style='medium', color='000000')
    border = Border(top=side, bottom=side, left=side, right=side)
    font   = Font(color="E1E1E1", bold=True, size=20)
    fill   = PatternFill("solid", fgColor="66547A")
    align  = Alignment(horizontal="center", vertical="center")

    headers = ["File A", "Column A", "File B", "Column B", "Rating"]
    for idx, title in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=idx, value=title)
        cell.font      = font
        cell.fill      = fill
        cell.alignment = align
        cell.border    = border
        ws.column_dimensions[cell.column_letter].width = (200 - 5) / 7

    body_font    = Font(color="FFFFFF", size=14)
    odd_fill     = PatternFill("solid", fgColor="403151")
    even_fill    = PatternFill("solid", fgColor="262626")
    left_align   = Alignment(horizontal="left",   vertical="center")
    center_align = Alignment(horizontal="center", vertical="center")

    for excel_row, rec in enumerate(relationships, start=2):
        data_idx = excel_row - 1
        fill     = odd_fill if (data_idx % 2 == 0) else even_fill

        ws.cell(row=excel_row, column=1, value=rec["file_a"]).font      = body_font
        ws.cell(row=excel_row, column=1).fill                            = fill
        ws.cell(row=excel_row, column=1).alignment                       = left_align

        ws.cell(row=excel_row, column=2, value=rec["column_a"]).font    = body_font
        ws.cell(row=excel_row, column=2).fill                            = fill
        ws.cell(row=excel_row, column=2).alignment                       = left_align

        ws.cell(row=excel_row, column=3, value=rec["file_b"]).font      = body_font
        ws.cell(row=excel_row, column=3).fill                            = fill
        ws.cell(row=excel_row, column=3).alignment                       = left_align

        ws.cell(row=excel_row, column=4, value=rec["column_b"]).font    = body_font
        ws.cell(row=excel_row, column=4).fill                            = fill
        ws.cell(row=excel_row, column=4).alignment                       = left_align

        ws.cell(row=excel_row, column=5, value=rec["rating"]).font      = body_font
        ws.cell(row=excel_row, column=5).fill                            = fill
        ws.cell(row=excel_row, column=5).alignment                       = center_align

    wb.save(report_path)
