# report.py
import pandas as pd
from openpyxl import load_workbook
import math

def format_time(seconds: float) -> str:
    if seconds < 60:
        return f"{int(seconds)} seconds"
    minutes = int(seconds // 60)
    sec = int(seconds % 60)
    return f"{minutes} min {sec} sec"

def pixels_to_excel_width(pixels: int) -> float:
    return (pixels - 5) / 7

def map_format_to_type(fmt: str) -> str:
    f = (fmt or "").lower()
    if any(tok in f for tok in ('y', 'd', 'm', 'h', 's')):
        return 'date'
    if any(tok in f for tok in ('0', '#', '%')):
        return 'number'
    return 'text'

def extract_column_types_from_excel(path: str) -> dict:
    from openpyxl import load_workbook as _load
    wb = _load(path, data_only=True)
    sheet = wb[wb.sheetnames[0]]
    headers = [cell.value for cell in sheet[1]]
    types = {}
    for idx, header in enumerate(headers):
        fmt = sheet.cell(row=2, column=idx+1).number_format or "General"
        types[header] = map_format_to_type(fmt)
    return types

def create_report(all_issues: dict,
                  time_stats: dict,
                  file_encodings: dict,
                  file_paths: dict,
                  output_path: str):
    with pd.ExcelWriter(output_path, engine='xlsxwriter') as writer:
        workbook = writer.book

        header_fmt = workbook.add_format({
            'bold': True,
            'font_color': '#E1E1E1',
            'bg_color': '#66547A',
            'font_size': 20,
            'align': 'center',
            'valign': 'vcenter',
            'top': 2,
            'bottom': 2,
            'left': 2,
            'right': 2
        })

        body_format_odd = workbook.add_format({
            'font_color': 'white', 'bg_color': '#403151',
            'font_size': 14, 'align': 'left', 'valign': 'vcenter'
        })
        body_format_even = workbook.add_format({
            'font_color': 'white', 'bg_color': '#262626',
            'font_size': 14, 'align': 'left', 'valign': 'vcenter'
        })

        center_format_odd = workbook.add_format({
            'font_color': 'white', 'bg_color': '#403151',
            'font_size': 14, 'align': 'center', 'valign': 'vcenter'
        })
        center_format_even = workbook.add_format({
            'font_color': 'white', 'bg_color': '#262626',
            'font_size': 14, 'align': 'center', 'valign': 'vcenter'
        })

        red_body_odd = workbook.add_format({
            'font_color': '#D23B3B', 'bg_color': '#403151',
            'font_size': 14, 'align': 'left', 'valign': 'vcenter'
        })
        red_body_even = workbook.add_format({
            'font_color': '#D23B3B', 'bg_color': '#262626',
            'font_size': 14, 'align': 'left', 'valign': 'vcenter'
        })

        green_body_odd = workbook.add_format({
            'font_color': '#00FF11', 'bg_color': '#403151',
            'font_size': 14, 'align': 'left', 'valign': 'vcenter'
        })
        green_body_even = workbook.add_format({
            'font_color': '#00FF11', 'bg_color': '#262626',
            'font_size': 14, 'align': 'left', 'valign': 'vcenter'
        })

        yellow_body_odd = workbook.add_format({
            'font_color': '#FFEE00', 'bg_color': '#403151',
            'font_size': 14, 'align': 'left', 'valign': 'vcenter'
        })
        yellow_body_even = workbook.add_format({
            'font_color': '#FFEE00', 'bg_color': '#262626',
            'font_size': 14, 'align': 'left', 'valign': 'vcenter'
        })

        orange_body_odd = workbook.add_format({
            'font_color': '#FF7700', 'bg_color': '#403151',
            'font_size': 14, 'align': 'left', 'valign': 'vcenter'
        })
        orange_body_even = workbook.add_format({
            'font_color': '#FF7700', 'bg_color': '#262626',
            'font_size': 14, 'align': 'left', 'valign': 'vcenter'
        })

        files_list = [key.split("(")[0].strip() for key in all_issues.keys()]
        encodings  = [file_encodings.get(f, '') for f in files_list]
        summary_df = pd.DataFrame({
            "Files Analyzed": files_list,
            "Encoding": encodings
        })
        summary_df.loc[0, "Analysis Time"] = format_time(time_stats.get("elapsed_s", 0))
        summary_df.to_excel(writer, sheet_name="Summary", index=False)
        ws_sum = writer.sheets["Summary"]

        for col_idx, col_name in enumerate(summary_df.columns):
            ws_sum.write(0, col_idx, col_name, header_fmt)
            ws_sum.set_column(col_idx, col_idx, pixels_to_excel_width(200))
        for row in range(1, len(summary_df) + 1):
            fmt = body_format_odd if row % 2 == 0 else body_format_even
            for col_idx in range(len(summary_df.columns)):
                val = summary_df.iloc[row - 1, col_idx]
                if val is None or (isinstance(val, float) and (math.isnan(val) or math.isinf(val))):
                    val = ""
                ws_sum.write(row, col_idx, val, fmt)

        column_headers = {
            "column": "Column",
            "type":   "Type",
            "issue":  "Error Type",
            "count":  "Count",
            "pct":    "Error Rate",
            "details":"Details",
            "rows":   "Rows"
        }

        for file_key, issues in all_issues.items():
            df_iss = pd.DataFrame(issues)[["column", "issue", "count", "pct", "details", "rows"]]

            basename = file_key.split()[0]
            path     = file_paths.get(basename, "")
            if path.lower().endswith('.xlsx'):
                col_types = extract_column_types_from_excel(path)
            else:
                col_types = {col: "text" for col in df_iss["column"].unique()}

            df_iss.insert(
                loc=df_iss.columns.get_loc("issue"),
                column="type",
                value=[col_types.get(col_name, "") for col_name in df_iss["column"]]
            )

            sheet_name = file_key[:31]
            df_iss.to_excel(writer, sheet_name=sheet_name, index=False)
            ws = writer.sheets[sheet_name]

            widths = {
                "column": 211, "type": 137, "issue": 394,
                "count": 117, "pct": 140, "details": 318,
                "rows": 580
            }
            for idx, col in enumerate(df_iss.columns):
                ws.set_column(idx, idx, pixels_to_excel_width(widths[col]))
                ws.write(0, idx, column_headers[col], header_fmt)

            centered_cols = {"count", "pct", "type"}
            for row in range(1, len(df_iss) + 1):
                is_even = (row % 2 == 0)
                for idx, col in enumerate(df_iss.columns):
                    val = df_iss.iloc[row - 1, idx]
                    if val is None or (isinstance(val, float) and (math.isnan(val) or math.isinf(val))):
                        val = "N/A"
                    
                    if col in centered_cols:
                        base_fmt = center_format_even if is_even else center_format_odd
                    else:
                        base_fmt = body_format_even if is_even else body_format_odd

                    if col == "issue":
                        if val == "Start date is after end date" or val=="Zero quantity with non-zero price"  or val=="Currency mismatch"  or val=="Male gender with female name"  or val=="Mixed Data Types"  or val=="Invalid Date Format":
                            fmt = red_body_even if is_even else red_body_odd
                        elif val == "Outliers"  or val=="Negative Values"  or val=="Zero Values"  or val=="Full Duplicate Rows"  or val=="Column Value Match"  or val=="There Are Some Columns Match":
                            fmt = green_body_even if is_even else green_body_odd
                        elif val == "Mostly Empty Column"  or val=="Time Repetition Error"  or val=="All values Missing On Column"  or val=="Missing Row":
                            fmt = yellow_body_even if is_even else yellow_body_odd
                        elif val == "Missing values"  or val=="There Are Symbols In Cells"  or val=="Found Unacceptable Keyword":
                            fmt = orange_body_even if is_even else orange_body_odd
                        else:
                            fmt = base_fmt
                    else:
                        fmt = base_fmt



                    ws.write(row, idx, val, fmt)
